import openpyxl
from openpyxl.styles import Alignment, Font
from datetime import date


wb = openpyxl.load_workbook('C:/Users/사용자명/Downloads/[PROD]대화 기록[20220000180637].xlsx')
ws = wb.active

ws.unmerge_cells('A1:H1') #row 1 병합 해제
ws.delete_rows(1) #row 1 삭제
ws.delete_cols(2) #column 2 삭제
ws.delete_cols(4,2) #column 4,5 삭제
ws.delete_cols(5,2) #column 5,6 삭제
ws.insert_cols(2) #column 2 삽입

columns = ws.iter_cols(min_col=4,max_col=4) #column 3에 있는 value를 column 2에 입력
for col in columns: 
    for cell in col:
        cell_new=ws.cell(row=cell.row, column=2, value=cell.value)

ws.delete_cols(4)

ws.column_dimensions['B'].width = 49.30 # 너비 조절
ws.column_dimensions['A'].width = 7.87 # 너비 조절
ws.row_dimensions[1].width = 30.00 # 높이 조절


for row in ws.iter_rows(min_col=2,max_col=2): # column 2 줄바꿈 적용
    for cell in row:
        cell.alignment=Alignment(wrap_text=True,vertical='top')

ws['B1'].font = Font(name='Calibri', size='11', bold=True)
ws['B1'].alignment = Alignment(vertical='center')

# column A merge
n=2
start_row =  2
start_column = 1
end_row = 2
end_column = 1
merge_max = ws.max_row + 1 


while end_row < merge_max :
    if ws['A'+str(n)].value == ws['A'+str(n+1)].value:
        end_row += 1
        n +=1

    elif ws['A'+str(n)].value != ws['A'+str(n+1)].value or ws['A'+str(n+1)].value == None:
        ws.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)

        start_row = end_row + 1
        end_row += 1
        n+= 1

columns = ws.iter_cols(min_col=1,max_col=1) #column A 가운데 정렬
for col in columns: 
    for cell in col:
        cell.alignment = Alignment(vertical='center')

columns = ws.iter_cols(min_col=4,max_col=5) #column D,E 가운데 정렬
for col in columns: 
    for cell in col:
        cell.alignment = Alignment(vertical='center')
      

today = date.today().isoformat().replace('-','')

wb.save(filename=f'C:/Users/사용자명/Downloads/{today}_음성봇_대화기록.xlsx')
wb.close




